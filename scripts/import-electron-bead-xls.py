#!/usr/bin/env python3
"""
Read Excel/WPS mapping workbook and emit mapping JSON + extension RGB overrides.
Run from repo root:
  python3 scripts/import-electron-bead-xls.py
  python3 scripts/import-electron-bead-xls.py /path/to/workbook.xlsx
"""
from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLS_PATH = Path.home() / "Downloads" / "电子拼豆.xlsx"
OUT_MAP = ROOT / "src" / "data" / "mappings"
OUT_OVR = ROOT / "src" / "data" / "palettes" / "extension-color-overrides.json"
OUT_EXTRA_META = ROOT / "src" / "data" / "palettes" / "non-mard-extension-keys.json"

MASTER_PATH = ROOT / "src" / "data" / "palettes" / "mard221.master.json"

KEY_RE = re.compile(r"^[A-Z]\d+$")
RGB_RE = re.compile(r"rgb\s*\(\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)\s*\)", re.I)
EXTRA_MASTER_PREFIX = "X"


def norm_cell(v) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    s = str(v).strip()
    return s if s else None


def parse_hex_cell(v) -> str | None:
    s = norm_cell(v)
    if not s:
        return None
    s = s.strip().lstrip("#")
    if len(s) != 6 or not re.fullmatch(r"[0-9a-fA-F]{6}", s):
        return None
    return f"#{s.lower()}"


def parse_rgb_cell(v) -> tuple[int, int, int] | None:
    s = norm_cell(v)
    if not s:
        return None
    m = RGB_RE.search(s)
    if not m:
        return None
    r, g, b = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if not all(0 <= x <= 255 for x in (r, g, b)):
        return None
    return r, g, b


def main() -> int:
    workbook_path = Path(sys.argv[1]).expanduser() if len(sys.argv) > 1 else XLS_PATH
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else "主表"

    if not workbook_path.exists():
        print(f"Missing file: {workbook_path}", file=sys.stderr)
        return 1

    with open(MASTER_PATH, encoding="utf-8") as f:
        master = json.load(f)
    master_ids = {row["id"] for row in master}

    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    candidates = [sheet_name, "主表", "色号卡"]
    sh = None
    for cand in candidates:
        if cand in wb.sheetnames:
            sh = wb[cand]
            break
    if sh is None:
        print(f"Workbook sheets={wb.sheetnames}; cannot find 主表/色号卡", file=sys.stderr)
        wb.close()
        return 1
    header = list(next(sh.iter_rows(min_row=1, max_row=1, values_only=True)))
    header_idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None and str(h).strip()}
    required = ["MARD", "COCO", "漫漫", "盼盼", "咪小窝", "HEX值", "RGB值"]
    if not all(k in header_idx for k in required):
        print("Unexpected header:", header, file=sys.stderr)
        wb.close()
        return 1

    brands = ["mard", "coco", "manman", "panpan", "mixiaowo"]
    col_idx = {
        "mard": header_idx["MARD"],
        "coco": header_idx["COCO"],
        "manman": header_idx["漫漫"],
        "panpan": header_idx["盼盼"],
        "mixiaowo": header_idx["咪小窝"],
    }
    hex_col = header_idx["HEX值"]
    rgb_col = header_idx["RGB值"]

    mappings: dict[str, dict[str, str | None]] = {b: {} for b in brands}
    overrides: dict[str, dict[str, int | str]] = {}
    skipped_zg: list[str] = []
    extra_non_mard_rows = 0
    extra_master_ids: list[str] = []

    extra_counter = 1

    def ingest_non_mard_row(
        row_vals,
        coco_col: int,
        manman_col: int,
        panpan_col: int,
        mixiaowo_col: int,
        hex_col_i: int,
        rgb_col_i: int,
    ) -> bool:
        nonlocal extra_counter, extra_non_mard_rows
        per_brand = {
            "mard": None,
            "coco": norm_cell(row_vals[coco_col] if len(row_vals) > coco_col else None),
            "manman": norm_cell(row_vals[manman_col] if len(row_vals) > manman_col else None),
            "panpan": norm_cell(row_vals[panpan_col] if len(row_vals) > panpan_col else None),
            "mixiaowo": norm_cell(row_vals[mixiaowo_col] if len(row_vals) > mixiaowo_col else None),
        }
        if not any(per_brand[b] for b in ["coco", "manman", "panpan", "mixiaowo"]):
            return False
        hex_v = parse_hex_cell(row_vals[hex_col_i] if len(row_vals) > hex_col_i else None)
        rgb_v = parse_rgb_cell(row_vals[rgb_col_i] if len(row_vals) > rgb_col_i else None)
        if not hex_v:
            return False

        key = f"{EXTRA_MASTER_PREFIX}{extra_counter}"
        while key in mappings["mard"]:
            extra_counter += 1
            key = f"{EXTRA_MASTER_PREFIX}{extra_counter}"
        extra_counter += 1

        mappings["mard"][key] = None
        mappings["coco"][key] = per_brand["coco"]
        mappings["manman"][key] = per_brand["manman"]
        mappings["panpan"][key] = per_brand["panpan"]
        mappings["mixiaowo"][key] = per_brand["mixiaowo"]

        if rgb_v:
            overrides[key] = {"hex": hex_v, "r": rgb_v[0], "g": rgb_v[1], "b": rgb_v[2]}
        else:
            hx = hex_v.lstrip("#")
            overrides[key] = {
                "hex": hex_v,
                "r": int(hx[0:2], 16),
                "g": int(hx[2:4], 16),
                "b": int(hx[4:6], 16),
            }
        extra_non_mard_rows += 1
        extra_master_ids.append(key)
        return True

    for row in sh.iter_rows(min_row=2, values_only=True):
        key_raw = row[0] if len(row) > 0 else None
        key = norm_cell(key_raw)
        if not key:
            ingest_non_mard_row(
                row,
                col_idx["coco"],
                col_idx["manman"],
                col_idx["panpan"],
                col_idx["mixiaowo"],
                hex_col,
                rgb_col,
            )
            continue
        if not KEY_RE.match(key):
            skipped_zg.append(key)
            continue

        hex_v = parse_hex_cell(row[hex_col] if len(row) > hex_col else None)
        rgb_v = parse_rgb_cell(row[rgb_col] if len(row) > rgb_col else None)
        # Workbook HEX/RGB should be the display truth for imported keys,
        # including master ids that also exist in mard221.master.json.
        if hex_v and rgb_v:
            overrides[key] = {"hex": hex_v, "r": rgb_v[0], "g": rgb_v[1], "b": rgb_v[2]}
        elif hex_v:
            hx = hex_v.lstrip("#")
            overrides[key] = {
                "hex": hex_v,
                "r": int(hx[0:2], 16),
                "g": int(hx[2:4], 16),
                "b": int(hx[4:6], 16),
            }

        for b in brands:
            ci = col_idx[b]
            code = norm_cell(row[ci] if len(row) > ci else None)
            if b == "mard":
                mappings[b][key] = code if code is not None else key
            else:
                mappings[b][key] = code

    # Also ingest rows that have no MARD key but do have other-brand code + HEX/RGB.
    extra_sheet = wb["无MARD保留项"] if "无MARD保留项" in wb.sheetnames else None
    if extra_sheet is not None:
        extra_head = list(next(extra_sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
        extra_idx = {str(h).strip(): i for i, h in enumerate(extra_head) if h is not None and str(h).strip()}
        if all(k in extra_idx for k in ["MARD", "COCO", "漫漫", "盼盼", "咪小窝", "HEX值", "RGB值"]):
            extra_hex_col = extra_idx["HEX值"]
            extra_rgb_col = extra_idx["RGB值"]
            for row in extra_sheet.iter_rows(min_row=2, values_only=True):
                mard_val = norm_cell(row[extra_idx["MARD"]] if len(row) > extra_idx["MARD"] else None)
                if mard_val:
                    continue
                ingest_non_mard_row(
                    row,
                    extra_idx["COCO"],
                    extra_idx["漫漫"],
                    extra_idx["盼盼"],
                    extra_idx["咪小窝"],
                    extra_hex_col,
                    extra_rgb_col,
                )

    def sort_map(m: dict[str, str | None]) -> dict[str, str | None]:
        return dict(sorted(m.items(), key=lambda x: x[0]))

    def dedupe_codes_per_brand(m: dict[str, str | None]) -> None:
        """BrandSettingsBundle requires unique non-null brandCode per brand."""
        by_code: dict[str, list[str]] = defaultdict(list)
        for k, v in m.items():
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            by_code[s].append(k)
        for code, keys in by_code.items():
            if len(keys) <= 1:
                continue
            for i, master_id in enumerate(sorted(keys)):
                if i == 0:
                    continue
                m[master_id] = f"{code}·{master_id}"

    for b in brands:
        dedupe_codes_per_brand(mappings[b])

    OUT_MAP.mkdir(parents=True, exist_ok=True)
    for b in brands:
        p = OUT_MAP / f"{b}.mapping.json"
        with open(p, "w", encoding="utf-8") as f:
            json.dump(sort_map(mappings[b]), f, ensure_ascii=False, indent=2)
            f.write("\n")

    with open(OUT_OVR, "w", encoding="utf-8") as f:
        json.dump(dict(sorted(overrides.items())), f, ensure_ascii=False, indent=2)
        f.write("\n")
    with open(OUT_EXTRA_META, "w", encoding="utf-8") as f:
        json.dump({"prefix": EXTRA_MASTER_PREFIX, "keys": extra_master_ids}, f, ensure_ascii=False, indent=2)
        f.write("\n")

    wb.close()
    print(
        f"Wrote {len(mappings['mard'])} keys per brand; "
        f"{len(overrides)} extension RGB overrides; "
        f"ingested non-MARD rows: {extra_non_mard_rows}."
    )
    if skipped_zg:
        print(f"Skipped invalid keys (not in ^[A-Z]\\d+$): {skipped_zg}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
